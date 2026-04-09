import openpyxl as xl
from fastapi import FastAPI
from model import Items,StockItems,ExpenseModel
from datetime import datetime
import random


app=FastAPI()

item_data=[["1","Advance",25,20],["2","Lite",25,20],["3","Gold Flake",12,10],["4","Double Switch",25,10]]




@app.post("/add_items")
def add_Items(item:Items=None):
    wb=xl.load_workbook("database.xlsx")
    sheet=wb["Items data"]
    if sheet.max_row==1:
        for i in item_data:
            sheet.append(i)
            wb.save("database.xlsx")
        wb.close()
        return{
            "message":"Items Updated "
        }
    for i,val in enumerate(sheet.iter_rows(values_only=True)):
        if i==0:
            continue
        if item.name.lower()==val[1].lower():
            return {
                "message":"Duplicated Items"
            }
        item_data.append([str(sheet.max_row),item.name,item.price,item.sticks])
    sheet.append([str(sheet.max_row),item.name,item.price,item.sticks])
    wb.save("database.xlsx")
    wb.close()
    return {
        "message":"Item Saved Successfully",
        "data":{
            "id":sheet.max_row-1,
            "name":item.name,
            "price":item.price,
            "sticks":item.sticks
        }
    }


@app.post("/add_stocks_data")
def add_stock_data(data_list: list[StockItems]):
    wb=xl.load_workbook("database.xlsx")
    opening_sheet=wb["Opening Stock"]
    stock_sheet=wb["Stock Data"]
    new_opening_data=[]
    opening_balance=0
    closing_balance=0
    for i,val in enumerate(opening_sheet.iter_rows(values_only=True)):
        if i==0: continue
        opening_sticks=(val[3]*val[5])+val[4]
        opening_balance +=opening_sticks*val[2]
        for data in data_list:
            if val[0]!=str(data.id):
                continue
            closing_sticks=(data.packs*data.stick_count)+data.sticks
            closing_balance +=closing_sticks*data.price
            remaining_sticks=opening_sticks-closing_sticks
            new_opening_data.append([data.id,data.name,data.price,data.packs,data.sticks,data.stick_count,datetime.now()])

    for data in new_opening_data:
        for i, val in enumerate(opening_sheet.iter_rows(values_only=True)):
            if i==0: continue
            if str(val[0])==str(data[0]):
                opening_sheet.cell(row=i+1,column=4).value=data[3]
                opening_sheet.cell(row=i+1,column=5).value=data[4]
                opening_sheet.cell(row=i+1,column=7).value=data[6]
                wb.save("database.xlsx")
        stock_sheet.append(data)
        wb.save("database.xlsx")
    credit_sheet=wb["Credit"]
    credit_sheet.append([datetime.now(),opening_balance,closing_balance,opening_balance-closing_balance])
    wb.save("database.xlsx")
    wb.close()
    return {
        "Opening Balance":opening_balance,
        "Closing Balance":closing_balance,
        "Profit Loss":opening_balance-closing_balance
    }


@app.post("/add_expense")
def add_expense(data_list:list[ExpenseModel]):
    wb=xl.load_workbook("database.xlsx")
    sheet=wb["Expense data"]
    try:
        for data in data_list:
            id=sheet.max_row
            sheet.append([str(id),data.name,data.type,data.amount,datetime.now()]) 
            wb.save("database.xlsx")
        wb.close()
        return {
            "Message":"Code Executed Successfully"
        }
    except  :
        return{
            "Message":"Something Went Wrong",
        }
