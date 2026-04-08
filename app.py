import openpyxl as xl

# Temporary File to Clear Worksheet

wb=xl.load_workbook("database.xlsx")

for i in wb.sheetnames:
    sheet = wb[i]
    sheet.delete_rows(1, sheet.max_row)
    if i=="Items data":
        sheet.append(["ID","Name","Price","Sticks"])
    if i=="Stock Data" or i=="Opening Stock":
        sheet.append(["Id","Name","Price","pack","Sticks","stick Count","Date"])
    if i=="Expense data":
        sheet.append(["id","Name","Type","Amount","Date"])
    if i=="Credit":
        sheet.append(["Date",	"Opening Balance",	"Closing Balance",	"Profit /Loss"])
    wb.save("database.xlsx")
    print(i)
wb.close()